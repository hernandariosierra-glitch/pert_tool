# app.py
import os, io, base64, json, tempfile, zipfile
from flask import Flask, render_template, request, jsonify, send_file
from model import Task, Project
from scheduler import calculate_schedule, assign_events, generate_table
from renderer import render_aoa

app = Flask(__name__)

OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "pert_outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def build_project(tasks_data):
    project = Project()
    for row in tasks_data:
        name = row.get("name", "").strip()
        if not name:
            continue
        duration     = int(row.get("duration", 0))
        preds_raw    = row.get("predecessors", "").strip()
        predecessors = [p.strip() for p in preds_raw.replace(",", " ").split() if p.strip()]
        project.add_task(Task(name, duration, predecessors))
    return project


def safe(name):
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/run", methods=["POST"])
def run_pert():
    data         = request.get_json()
    project_name = (data.get("project_name") or "Proyecto_PERT").strip()
    tasks_data   = data.get("tasks", [])

    if not tasks_data:
        return jsonify({"error": "No hay tareas ingresadas."}), 400

    names = [r.get("name", "").strip() for r in tasks_data if r.get("name", "").strip()]
    if len(names) != len(set(names)):
        return jsonify({"error": "Hay nombres de actividad duplicados."}), 400

    try:
        project  = build_project(tasks_data)
        duration, critical_path = calculate_schedule(project)
        assign_events(project)
    except Exception as e:
        return jsonify({"error": f"Error en el calculo: {e}"}), 500

    sname    = safe(project_name)
    png_path = os.path.join(OUTPUT_DIR, f"{sname}_pert")

    try:
        render_aoa(project, filename=png_path)
    except Exception as e:
        return jsonify({"error": f"Error al generar el diagrama: {e}"}), 500

    with open(png_path + ".png", "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode()

    table = [
        {
            "name":     r["Tarea"],
            "duration": r["Duración"],
            "es": r["ES"], "ef": r["EF"],
            "ls": r["LS"], "lf": r["LF"],
            "slack":    r["Holgura"],
            "critical": r["Holgura"] == 0,
        }
        for r in generate_table(project)
    ]

    saved = {
        "project_name":  project_name,
        "duration":      duration,
        "critical_path": " -> ".join(critical_path),
        "table":         table,
        "tasks":         [
            {"name": r.get("name",""), "duration": r.get("duration",0),
             "predecessors": r.get("predecessors","")}
            for r in tasks_data if r.get("name","").strip()
        ],
    }
    with open(os.path.join(OUTPUT_DIR, f"{sname}_data.json"), "w") as f:
        json.dump(saved, f)

    return jsonify({
        "ok":            True,
        "duration":      duration,
        "critical_path": " -> ".join(critical_path),
        "image_b64":     img_b64,
        "table":         table,
        "safe_name":     sname,
    })


@app.route("/api/download/<sname>/<fmt>")
def download(sname, fmt):
    png_path  = os.path.join(OUTPUT_DIR, f"{sname}_pert.png")
    json_path = os.path.join(OUTPUT_DIR, f"{sname}_data.json")

    if not os.path.exists(png_path):
        return "Archivo no encontrado. Vuelva a generar el diagrama.", 404

    if fmt == "png":
        return send_file(png_path, mimetype="image/png",
                         as_attachment=True,
                         download_name=f"{sname}_pert.png")

    if fmt == "xlsx":
        xlsx_path = os.path.join(OUTPUT_DIR, f"{sname}_resultados.xlsx")
        _build_xlsx(json_path, xlsx_path)
        return send_file(xlsx_path,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name=f"{sname}_resultados.xlsx")

    if fmt == "zip":
        xlsx_path = os.path.join(OUTPUT_DIR, f"{sname}_resultados.xlsx")
        _build_xlsx(json_path, xlsx_path)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.write(png_path,  f"{sname}_pert.png")
            zf.write(xlsx_path, f"{sname}_resultados.xlsx")
        buf.seek(0)
        return send_file(buf, mimetype="application/zip",
                         as_attachment=True,
                         download_name=f"{sname}_pert.zip")

    return "Formato no soportado.", 400


def _build_xlsx(json_path, xlsx_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with open(json_path) as f:
        saved = json.load(f)

    wb        = openpyxl.Workbook()
    hdr_fill  = PatternFill("solid", fgColor="1A3A5C")
    crit_fill = PatternFill("solid", fgColor="FFE5E5")
    alt_fill  = PatternFill("solid", fgColor="EEF4FF")
    wht_fill  = PatternFill("solid", fgColor="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def style_header(ws, headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin

    def style_cell(cell, fill):
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    ws1 = wb.active
    ws1.title = "Resultados"
    h1 = ["Actividad", "Duracion", "ES", "EF", "LS", "LF", "Holgura", "Critica?"]
    style_header(ws1, h1)

    for i, row in enumerate(saved["table"], start=2):
        ws1.append([row["name"], row["duration"],
                    row["es"], row["ef"], row["ls"], row["lf"],
                    row["slack"], "Si" if row["critical"] else "No"])
        fill = crit_fill if row["critical"] else (alt_fill if i % 2 == 0 else wht_fill)
        for col in range(1, len(h1) + 1):
            style_cell(ws1.cell(row=i, column=col), fill)

    ws1.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws1.column_dimensions[col].width = 10

    last = len(saved["table"]) + 3
    ws1.cell(row=last,     column=1, value="Duracion total:").font = Font(bold=True)
    ws1.cell(row=last,     column=2, value=saved["duration"])
    ws1.cell(row=last + 1, column=1, value="Ruta critica:").font   = Font(bold=True)
    ws1.cell(row=last + 1, column=2, value=saved["critical_path"])

    ws2 = wb.create_sheet("Datos ingresados")
    h2 = ["ID", "Duracion", "Predecesores"]
    style_header(ws2, h2)
    for i, task in enumerate(saved["tasks"], start=2):
        ws2.append([task["name"], task["duration"], task["predecessors"]])
        fill = alt_fill if i % 2 == 0 else wht_fill
        for col in range(1, 4):
            style_cell(ws2.cell(row=i, column=col), fill)
    for col in "ABC":
        ws2.column_dimensions[col].width = 18

    wb.save(xlsx_path)


if __name__ == "__main__":
    app.run(debug=True, port=5000)